from io import BytesIO
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles import Alignment, PatternFill

from sqlalchemy import func

from grimm import logger, db
from grimm.models.activity import Duty, Gift
from grimm.models.activity import ActivityParticipant, PickupImpaired, Activity, PickupVolunteer
from grimm.models.admin import User
from grimm.utils import misctools, constants, smstools


def activity_converter(activity, openid=0):
    query = {}
    logger.info('Convert activity dto.')
    query["id"] = activity["id"]
    query["adminId"] = activity["approver"]
    query["title"] = activity["title"]
    query["location"] = activity["location"]
    query["sign_in_radius"] = activity["sign_in_radius"]
    query["sign_in_token"] = activity["sign_in_token"]
    query["start_time"] = activity["start_time"].strftime("%Y-%m-%dT%H:%M:%S")
    query["end_time"] = activity["end_time"].strftime("%Y-%m-%dT%H:%M:%S")
    query["duration"] = misctools.calc_duration(activity["start_time"], activity["end_time"])
    query["content"] = activity["content"]
    query["notice"] = activity["notice"]
    query["others"] = activity["others"]
    query["tag"] = ','.join([constants.TAG_LIST[int(tid)] for tid in activity["tag_ids"].split(',') if tid or int(tid) in range(6)]) \
        if activity["tag_ids"] else ''
    if openid == 0:
        participant = ActivityParticipant.query.filter(ActivityParticipant.activity_id == activity["id"]).all()
        query["share"] = sum([int(part.share) for part in participant if (part and part.share)])
    else:
        participant = ActivityParticipant.query.filter(ActivityParticipant.activity_id == activity["id"],
                                                       ActivityParticipant.participant_openid == openid).first()
        if participant:
            query["share"] = int(participant.share) if participant and participant.share else 0
        else:
            query["share"] = 0
    query["interested"] = ActivityParticipant.query. \
        filter(ActivityParticipant.activity_id == activity["id"],
               ActivityParticipant.interested == 1).count() if openid == 0 \
        else ActivityParticipant.query.filter(ActivityParticipant.activity_id == activity["id"],
                                              ActivityParticipant.participant_openid == openid,
                                              ActivityParticipant.interested == 1).count()
    query["thumbs_up"] = ActivityParticipant.query. \
        filter(ActivityParticipant.activity_id == activity["id"],
               ActivityParticipant.thumbs_up == 1).count() if openid == 0 \
        else ActivityParticipant.query.filter(ActivityParticipant.activity_id == activity["id"],
                                              ActivityParticipant.participant_openid == openid,
                                              ActivityParticipant.thumbs_up == 1).count()

    query["registered"] = ActivityParticipant.query. \
        filter(ActivityParticipant.activity_id == activity["id"],
               ActivityParticipant.current_state != None).count() if openid == 0 else \
        ActivityParticipant.query.filter(ActivityParticipant.activity_id == activity["id"],
                                         ActivityParticipant.participant_openid == openid,
                                         ActivityParticipant.current_state != None).count()
    query["registered_volunteer"] = db.session.query(ActivityParticipant, User). \
        filter(ActivityParticipant.activity_id == activity["id"],
               ActivityParticipant.current_state.in_(('Registered', 'signed_up', 'signed_off'))). \
        filter(User.role == 0). \
        filter(ActivityParticipant.participant_openid == User.openid).count()
    query["registered_impaired"] = db.session.query(ActivityParticipant, User). \
        filter(ActivityParticipant.activity_id == activity["id"],
               ActivityParticipant.current_state.in_(('Registered', 'signed_up', 'signed_off'))). \
        filter(User.role == 1). \
        filter(ActivityParticipant.participant_openid == User.openid).count()
    query["volunteer_capacity"] = activity["volunteer_capacity"]
    query["is_volunteer_limited"] = (
        True
        if (
                activity["volunteer_capacity"] is not None
                and activity["volunteer_capacity"] > 0
        )
        else False
    )
    query["vision_impaired_capacity"] = activity["vision_impaired_capacity"]
    query["is_impaired_limited"] = (
        True
        if (
                activity["vision_impaired_capacity"] is not None
                and activity["vision_impaired_capacity"] > 0
        )
        else False
    )
    query["volunteer_job_title"] = activity["volunteer_job_title"]
    query["volunteer_job_content"] = activity["volunteer_job_content"]
    query["activity_fee"] = activity["activity_fee"]
    query["is_fee_needed"] = (
        True
        if (activity["activity_fee"] is not None and activity["activity_fee"] > 0)
        else False
    )
    query["activity_them_pic_name"] = activity['theme_pic_name']
    return query


def sort_by_time(activities_info, filter_time):
    if filter_time == "all":
        res_info = [
                activity
                for activity in activities_info
                if datetime.today() - timedelta(days=365) < activity["end_time"]
            ]
    elif filter_time == "latest":
        res_info = [
                activity
                for activity in activities_info
                if datetime.today() < activity["end_time"]
            ]
    elif filter_time == "weekends":
        res_info = [
            activity
            for activity in activities_info
            if should_append_by_weekends(activity)
        ]
    elif filter_time == "recents":
        res_info = [
            activity
            for activity in activities_info
            if should_append_by_recents(activity)
        ]
    else:
        res_info = [
            activity
            for activity in activities_info
            if should_append_by_time_span(activity, filter_time)
        ]

    return sorted(res_info, reverse=True,
            key=lambda activity: activity["start_time"])


def should_append_by_time_span(activity, filter_time):
    filter_start = datetime.strptime(filter_time.split(" - ")[0], "%Y-%m-%d")
    filter_end = datetime.strptime(filter_time.split(" - ")[1], "%Y-%m-%d") + timedelta(days=1)
    start = activity["start_time"]
    end = activity["end_time"]
    if filter_end < start or filter_start > end:
        return False
    return True


def should_append_by_tag(activity, target_tag_list):
    if not activity:
        return False
    if target_tag_list == "all":
        return True
    if activity["tag_ids"] is not None:
        current_tag_list = activity["tag_ids"].split(",")
        for target_tag_id in target_tag_list.split(","):
            if target_tag_id in current_tag_list:
                return True
    return False


def should_append_by_weekends(activity):
    today = datetime.today()
    end = activity["end_time"]
    if today > end:
        return False
    start = activity["start_time"] if activity["start_time"] > today else today
    while start < end:
        if start.weekday() >= 5:
            return True
        start += timedelta(days=1)
    return False


def should_append_by_recents(activity):
    filter_start = datetime.today()
    filter_end = filter_start + timedelta(days=7)
    start = activity["start_time"]
    end = activity["end_time"]
    if filter_end < start or filter_start > end:
        return False
    return True


def user_cancel_activity(openid, activity_id):
    """ Volunteer cancel participation from wechat-end, should notice impaired or volunteer asap."""
    user_info = User.query.filter(User.openid == openid).first()
    activity_info = Activity.query.filter(Activity.id == activity_id).first()
    logger.info('User %s cancel activity %s, should remove the '
                'binding and give some notifications.' % (user_info.name, activity_info.title))
    if user_info.role == 0:
        logger.info('volunteer %s cancel activity %s.' % (user_info.name, activity_info.title))
        pick_info = db.session.query(PickupVolunteer).\
            filter(PickupVolunteer.openid == openid,
                   PickupVolunteer.activity_id == activity_id).first()
        if pick_info:
            logger.info('Query pickup list for volunteer %s' % user_info.name)
            pick_list = PickupImpaired.query. \
                filter(PickupImpaired.pickup_volunteer_openid == openid,
                       PickupImpaired.activity_id == activity_id).all()
            if pick_list:
                for pick in pick_list:
                    impaired_openid = pick.openid
                    pick_impaired = db.session.query(PickupImpaired). \
                        filter(PickupImpaired.activity_id == activity_id,
                               PickupImpaired.openid == impaired_openid).first()
                    logger.info('Clear pickup info and notice impaired %s' % pick_impaired.name)
                    pick_impaired.pickup_method = None
                    pick_impaired.pickup_volunteer_openid = None
                    db.session.commit()
                    kwargs = {
                        'impaired_name': pick_impaired.name,
                        'volunteer_name': pick_info.name,
                        'volunteer_phone': user_info.phone
                    }
                    impaired_user_info = User.query.filter(User.openid == impaired_openid).first()
                    phone_number_list = [impaired_user_info.phone]
                    template_id = constants.TEMPLATE_CODES['VOLUNTEER_CANCEL_ACTIVITY']
                    smstools.send_short_message(phone_number_list, template_id, **kwargs)
                logger.info('all impaired notice over.')
            db.session.delete(pick_info)
            db.session.commit()
    else:
        logger.info('Impaired %s cancel activity %s.' % (user_info.name, activity_info.title))
        pickup_impaired = db.session.query(PickupImpaired). \
            filter(PickupImpaired.openid == openid,
                   PickupImpaired.activity_id == activity_id).first()
        if pickup_impaired:
            logger.info('Have volunteer pickup current impaired? if yes, need notice volunteer.')
            if pickup_impaired.pickup_volunteer_openid and pickup_impaired.pickup_method:
                volunteer_user_info = User.query.filter(User.openid == pickup_impaired.pickup_volunteer_openid).first()
                logger.info("Current impaired volunteer is %s" % volunteer_user_info.name)
                kwargs = {
                    'impaired_name': pickup_impaired.name,
                    'volunteer_name': volunteer_user_info.name,
                    'impaired_phone': user_info.phone
                }
                phone_number_list = [volunteer_user_info.phone]
                template_id = constants.TEMPLATE_CODES['IMPAIRED_CANCEL_ACTIVITY']
                smstools.send_short_message(phone_number_list, template_id, **kwargs)
            db.session.delete(pickup_impaired)
            db.session.commit()
    logger.info('Volunteer or impaired %s cancel activity success.' % user_info.name)


def volunteer_pickup_impaired(volunteer_openid, impaired_openid, pickup_method):
    """ volunteer choose pickup impaired and detail pickup method """
    logger.info('Volunteer choose pickup method. need to notice impaired.')
    volunteer_user_info = User.query.filter(User.openid == volunteer_openid).first()
    impaired_user_info = User.query.filter(User.openid == impaired_openid).first()
    kwargs = {'impaired_name': impaired_user_info.name,
              'volunteer_name': volunteer_user_info.name,
              'volunteer_phone': volunteer_user_info.phone}
    phone_number_list = [impaired_user_info.phone]
    if pickup_method:
        logger.info('Volunteer %s will pickup %s' % (volunteer_user_info.name, impaired_user_info.name))
        template_id = constants.TEMPLATE_CODES['VOLUNTEER_PICKUP']
        smstools.send_short_message(phone_number_list, template_id, **kwargs)
    else:
        logger.info('Volunteer %s not pickup %s' % (volunteer_user_info.name, impaired_user_info.name))
        template_id = constants.TEMPLATE_CODES['VOLUNTEER_CANCEL_PICKUP']
        smstools.send_short_message(phone_number_list, template_id, **kwargs)

light_grey = PatternFill(start_color="DBDBDB", end_color="DBDBDB", fill_type = "solid")
dark_grey = PatternFill(start_color="9C9C9C", end_color="9C9C9C", fill_type = "solid")
yellow = PatternFill(start_color="FFF200", end_color="FFF200", fill_type = "solid")
center_aligned = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

def bold_with_size(size):
    return Font(bold=True, size=size)

bold_12 = bold_with_size(12)
bold_11 = bold_with_size(11)

def form_sign(activity):
    def _write_sheet(ws, users):
        ws.merge_cells('A1:H1')
        ws['A1'].fill = dark_grey
        ws['A1'].font = bold_with_size(20)
        ws['A1'].alignment = center_aligned
        ws['A1'].value = f'{ws.title} 签到/签收表'
        for cell in ws['A1':'H1'][0]:
            cell.border = thin_border

        project_name = activity.project.name
        if activity.project_seq:
            project_name += f' 第({activity.project_seq})期'
        ws.append(['', f'项目名称：{project_name}'])
        ws.merge_cells('B2:H2')

        ws.append(['',
            f'活动日期：{activity.start_date}'])
        ws.merge_cells('B3:H3')

        ws.append(['', f'活动主题：{activity.title}'])
        ws.merge_cells('B4:H4')

        ws.append(['序号', '姓名', '电话', '签名', '身份证', '衣服领取', '物品领取', '备注'])

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 10
        for col in 'CDEFGH':
            ws.column_dimensions[col].width = 20

        for row in ws['A2':'H5']:
            for cell in row:
                cell.font = bold_11
                cell.border = thin_border

        for cell in ws['A5':'H5'][0]:
            cell.fill = light_grey

        for idx, volunteer in enumerate(users):
            ws.append([
                idx+1,
                volunteer.name,
                volunteer.phone,
                '',  # column for manual input of signature
                volunteer.idcard,
                '', '', # columns for manual input of receiving items
                # TODO should be ActivityParticipant.remark?
                volunteer.remark])

        user_total = len(users)
        summary_row_idx = 36
        if user_total > 30:
            summary_row_idx = user_total + 6

        ws[f'A{summary_row_idx}'].value = '总计'
        ws[f'B{summary_row_idx}'].value = f'活动人数：  {user_total}  参加人数：    未参加人数：'
        ws.merge_cells(f'B{summary_row_idx}:H{summary_row_idx}')
        ws[f'A{summary_row_idx}'].font = bold_11
        ws[f'B{summary_row_idx}'].font = bold_with_size(14)
        ws[f'B{summary_row_idx}'].alignment = center_aligned
        for cell in ws[f'A{summary_row_idx}':f'H{summary_row_idx}'][0]:
            cell.border = thin_border

    wb = Workbook()
    ws1 = wb.create_sheet(title='志愿')
    ws2 = wb.create_sheet(title='视障')

    _write_sheet(ws1, activity.volunteers)
    _write_sheet(ws2, activity.impaireds)

    # created two new sheet, so remove the first default sheet
    # which should be empty
    wb.remove(wb[wb.sheetnames[0]])

    wb.save('test.xlsx')

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return stream

def join_with_chinese_comma(str_list):
    return '、'.join(str_list)

def form_duty_summary(activities):
    all_duties = db.session.query(Duty).all()
    all_duties.sort(key=lambda x: x.seq)
    duty_names = [x.name for x in all_duties]

    wb = Workbook()
    ws = wb.active
    title_row = ['活动序号', '日期']
    title_row.extend(duty_names)
    ws.append(title_row)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    for col in 'CDEFGH':
        ws.column_dimensions[col].width = 50

    for cell in ws['A1':'H1'][0]:
        cell.fill = light_grey
        cell.font = bold_12
        cell.border = thin_border

    for idx, activity in enumerate(activities):
        one = [idx+1, activity.start_date]
        for name in duty_names:
            one.append(join_with_chinese_comma([info.user.name for info \
                in activity.participate_infos if info.duty_name == name]))
        ws.append(one)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return stream

def form_info_summary(activities):
    all_gifts = db.session.query(Gift).all()
    all_gifts.sort(key=lambda x: x.seq)
    gift_names = [x.name for x in all_gifts]

    wb = Workbook()
    ws = wb.active
    ws.merge_cells('A1:L1')
    ws['A1'].font = bold_with_size(20)
    ws['A1'].alignment = center_aligned
    ws['A1'].value = '活动汇总信息'
    ws['A1'].border = thin_border

    title_row = ['活动序号', '日期', '地点', '孩子数', '志愿者人数', '视障者人数', '总人数']
    for name in gift_names:
        title_row.append(f'{name}发件数')
    title_row.append('备注')
    ws.append(title_row)
    for cell in ws['A2':'L2'][0]:
        cell.fill = yellow
        cell.font = bold_12
        cell.border = thin_border

    for idx, activity in enumerate(activities):
        one = [idx+1, activity.start_date, activity.location,
                activity.children_count, len(activity.volunteers),
                len(activity.impaireds),
                len(activity.volunteers) + len(activity.impaireds),
            ]

        gift_count = {}
        for info in activity.participate_infos:
            if info.gifts:
                for _id in info.gifts:
                    gift_count.setdefault(_id, 0)
                    gift_count[_id] += info.gifts[_id]

        for g in all_gifts:
            one.append(gift_count.get(str(g.id), 0))

        one.append(activity.remark)
        ws.append(one)

    ws.column_dimensions['A'].width = 5
    for col in 'BDEFGHIJK':
        ws.column_dimensions[col].width = 10
    for col in 'CL':
        ws.column_dimensions[col].width = 20

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return stream
