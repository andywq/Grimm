import io
import os
import json
import pathlib
import unittest
from unittest import mock
from datetime import datetime

from openpyxl import Workbook, load_workbook
from sqlalchemy import func as sql_func
from werkzeug.datastructures import BytesIO

os.environ['FLASK_ENV'] = 'test'

from config import BASE_DIR
from grimm import db
from grimm.models.activity import Activity, ActivityParticipant
from grimm.models.admin import User

from .base import post_json
from .base import ActivityCase

# GET "/activities"
class TestGetActivities(ActivityCase):
    def test_get_all_activities(self) : #, mock_query, mock_serialize):
        response = self.client.get('/activities')
        self.assertEqual(response.status_code, 200)
        data = response.json
        self.assertIsInstance(data, list)
        # initialized 2 activities in setUp
        self.assertEqual(2, len(data))
        self.assertGreater(data[0]['start_time'],
            data[1]['start_time'])
        self.assertEqual(data[0]['project_id'], 1)
        self.assertEqual(data[0]['project_name'],
                self.default_projects[0]['name'])

    def test_get_activities_by_keyword(self):
        response = self.client.get('/activities?keyword={}'.format(
            self.default_activities[0]['title']))
        self.assertEqual(response.status_code, 200)
        self.assertIsInstance(response.json, list)
        self.assertEqual(1, len(response.json))
        # TODO pick a few attrs, for some attributes names are
        # not the same
        for k in ('title', 'location', 'content'):
            self.assertEqual(response.json[0][k],
                self.default_activities[0][k])

    def test_get_activities_by_tags(self):
        response = self.client.get('/activities?tags=1')
        self.assertEqual(response.status_code, 200)
        self.assertIsInstance(response.json, list)
        self.assertEqual(2, len(response.json))

    def test_get_activities_by_not_exists_tags(self):
        response = self.client.get('/activities?tags=null')
        self.assertEqual(response.status_code, 200)
        self.assertIsInstance(response.json, list)
        self.assertEqual(0, len(response.json))

    def test_get_activities_by_time(self):
        response = self.client.get('/activities?time=all')
        self.assertEqual(response.status_code, 200)
        self.assertIsInstance(response.json, list)

# POST "/activity"
class TestNewActivityResource(ActivityCase):
    test_data = {
        'title': 'activity_title_3',
        'adminId': 1,
        'location': 'activity_localtion_3',
        'content': 'activity_content_3',
        'notice': 'activity_notice_3',
        'others': 'activity_others_3',
        'activity_them_pic_name': [{
            'url': '/activity_them_3.png'
        }],
        'tag': '运动,学习',
        'volunteer_job_title': 'activity_job_title_3',
        'volunteer_job_content': 'activity_job_content_3',
        'sign_in_token': 's_token',
        'start_time': datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        'end_time': datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        'activity_fee': 0,
        'admin_raiser': 1,
        'volunteer_capacity': 1,
        'vision_impaired_capacity': 1,
        'sign_in_radius': 1,
    }

    @mock.patch('grimm.utils.areautils.address_to_coordinate')
    def test_post_new_activity_success(self, mock_address_to_coordinate):
        # Mock the address_to_coordinate function to return a successful status and coordinates
        mock_address_to_coordinate.return_value = (True, {'lat': 123.45, 'lng': 678.90})

        with self.app.app_context():
            max_project_seq = db.session.query(sql_func.max(Activity.project_seq)).filter_by(project_id=1).scalar()
            if max_project_seq is None:
                max_project_seq = 1

        response = post_json(self.client, '/activity', self.test_data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json, {"status": "success"})

        with self.app.app_context():
            activity = db.session.query(Activity).filter_by(title=self.test_data['title']).first()
            self.assertIsNotNone(activity)
            self.assertEqual(activity.project_id, 1)
            self.assertEqual(activity.project_seq, max_project_seq + 1)

    @mock.patch('grimm.utils.areautils.address_to_coordinate')
    def test_post_new_activity_failure_address(self, mock_address_to_coordinate):
        # Mock the address_to_coordinate function to return a failure status
        mock_address_to_coordinate.return_value = (False, {})

        response = post_json(self.client, '/activity', self.test_data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json, {"status": "failure"})

# GET "/activity/<int:activity_id>"
class TestOneActivity(ActivityCase):
    def test_get_activity_success(self):
        response = self.client.get('/activity/1')
        self.assertEqual(response.status_code, 200)
        self.assertIn('activity_them_pic_name', response.json)
        self.assertEqual(response.json['title'], self.default_activities[0]['title'])
        self.assertNotEqual(response.json['title'], self.default_activities[1]['title'])

    def test_get_activity_not_found(self):
        response = self.client.get('/activity/999')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json, {"status": "failure", "message": "未知活动ID"})

# POST "/activity/<int:activity_id>"
class TestActivityOperatePost(ActivityCase):
    @mock.patch('grimm.utils.areautils.address_to_coordinate')
    def test_update_activity_success(self, mock_address_to_coordinate):
        # Mock the address_to_coordinate function to return a successful status and coordinates
        mock_address_to_coordinate.return_value = (True, {'lat': 123.45, 'lng': 678.90})
        new_info = {
            'activity_them_pic_name': [{'url': 'new_pic.png'}],
            'adminId': 1,
            'title': 'new_title',
            'location': 'new_location',
            'sign_in_radius': 1,
            'sign_in_token': 'new_s_i_t',
            'start_time': datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            'end_time': datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            'content': 'new_content',
            'notice': 'new_notice',
            'others': 'new_others',
            'tag': '',
            'volunteer_job_title': 'new_activity_job_title',
            'volunteer_job_content': 'new_activity_job_content',
            'activity_fee': 0,
            'volunteer_capacity': 1,
            'vision_impaired_capacity': 1,
            'sign_in_radius': 1,
        }
        response = post_json(self.client, '/activity/1', data=new_info)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json, {"status": "success"})

        with self.app.app_context():
            activity = db.session.query(Activity).filter_by(id=1).first()
            self.assertEqual(activity.title, new_info['title'])
            self.assertEqual(activity.location, new_info['location'])

    def test_update_activity_invalid_id(self):
        new_info = {
            'activity_them_pic_name': [{'url': 'new_pic.png'}],
            'adminId': 'admin_123',
            'title': 'New Title',
            'location': '123 Main St',
        }
        response = self.client.post('/activity/999', data=json.dumps(new_info), content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json, {"status": "failure", "message": "无效活动 ID"})

    def test_update_activity_no_theme_picture(self):
        new_info = {
            'activity_them_pic_name': [],
            'adminId': 'admin_123',
            'title': 'New Title',
            'location': '123 Main St',
            # ... include other necessary fields ...
        }
        response = self.client.post('/activity/1', data=json.dumps(new_info), content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json, {"status": "failure", "message": "请上传活动主题图片"})

# DELETE "/activity/<int:activity_id>"
class TestActivityOperateDelete(ActivityCase):
    def test_delete_activity_success(self):
        response = self.client.delete('/activity/1')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json, {"status": "success", "message": "活动删除成功！"})
        with self.app.app_context():
            activity = db.session.query(Activity).filter_by(id=1).first()
            self.assertIsNone(activity)

    def test_delete_activity_with_participants(self):
        # Add a participant to the activity to simulate a delete failure due to existing participants
        with self.app.app_context():
            participant = ActivityParticipant(activity_id=1, participant_openid=self.default_volunteer_attrs['openid'])
            db.session.add(participant)
            db.session.commit()

        response = self.client.delete('/activity/1')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json, {"status": "failure", "message": "已有用户操作过该活动，不能删除！"})

# GET "/activity/themePic"
class TestActivityThemePicGet(ActivityCase):
    def test_get_theme_pic_by_activity_id(self):
        filename = self.default_activities[0]['theme_pic_name']
        image = os.path.join(BASE_DIR, "static/activity_theme_pictures", filename)
        pathlib.Path(image).touch()

        try:
            with self.app.test_request_context():
                response = self.client.get('/activity/themePic?activity_id=1')
                self.assertEqual(response.status_code, 200)
                self.assertTrue('image/jpeg' in response.content_type)
        finally:
            os.remove(image)

    def test_get_theme_pic_by_file_name(self):
        filename = self.default_activities[0]['theme_pic_name']
        image = os.path.join(BASE_DIR, "static/activity_theme_pictures", filename)
        pathlib.Path(image).touch()

        try:
            with self.app.test_request_context():
                response = self.client.get('/activity/themePic?activity_them_pic_name={}'.format(filename))
                self.assertEqual(response.status_code, 200)
                self.assertTrue('image/jpeg' in response.content_type)
        finally:
            os.remove(image)

    def test_get_theme_pic_no_params(self):
        with self.app.test_request_context():
            response = self.client.get('/activity/themePic')
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json, {"status": 'failure', "message": "Please input activity id or file name."})

    def test_get_theme_pic_activity_not_exist(self):
        with self.app.test_request_context():
            response = self.client.get('/activity/themePic?activity_id=999')
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json, {"status": 'failure', "message": "Activity not exists."})

# POST "/activity/themePic"
class TestActivityThemePicPost(ActivityCase):
    def test_upload_theme_pic(self):
        # Create a test file
        test_file_name = 'test-upload.jpg'
        test_file_content = b'This is a test file.'
        test_file = (BytesIO(test_file_content), test_file_name)
        data = {
            'activity_them_pic_name': test_file
        }
        response = self.client.post('/activity/themePic', data=data, content_type='multipart/form-data')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json['status'], 1)
        self.assertIn('fileName', response.json)
        # Check if file exists in the directory
        uploaded_file_name = response.json['fileName']
        full_path = os.path.join(BASE_DIR, "static/activity_theme_pictures/", uploaded_file_name)
        try:
            self.assertTrue(os.path.exists(full_path))
        finally:
            os.remove(full_path)

    def test_upload_no_file(self):
        response = self.client.post('/activity/themePic', content_type='multipart/form-data')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json, {"status": 0, "message": 'No file found'})

# GET "/activityRegistration/<int:activity_id>"
class TestActivityRegistrationGet(ActivityCase):
    def test_get_activity_registration(self):
        openid = self.default_volunteer_attrs['openid']
        gifts = {'1':2, '2':1}
        duties = [1,2]
        with self.app.app_context():
            participant = ActivityParticipant(activity_id=1, participant_openid=openid, gifts=gifts, duties=duties)
            db.session.add(participant)
            db.session.commit()

        response = self.client.get('/activityRegistration/1')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json['status'], 'success')
        self.assertIsInstance(response.json['users'], list)
        self.assertEqual(len(response.json['users']), 1)
        self.assertEqual(response.json['users'][0]['openid'], openid)
        self.assertEqual(response.json['users'][0]['name'], self.default_volunteer_attrs['name'])
        self.assertEqual(response.json['users'][0]['gifts'], gifts)
        self.assertEqual(response.json['users'][0]['duties'], duties)

    def test_get_activity_registration_no_activity(self):
        response = self.client.get('/activityRegistration/999')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json['status'], 'failure')
        self.assertEqual(response.json['message'], '无效活动 ID')

# POST "/activityRegistration/<int:activity_id>"
class TestActivityRegistrationPost(ActivityCase):
    def test_post_activity_registration_no_activity(self):
        response = self.client.post('/activityRegistration/999?openid=test_openid')
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertEqual(data['status'], 'failure')
        self.assertEqual(data['message'], '无效活动 ID')

    def test_post_activity_registration_not_signed_up(self):
        response = self.client.post('/activityRegistration/1?openid={}'.format(self.default_volunteer_attrs['openid']))
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertEqual(data['status'], 'failure')
        self.assertEqual(data['message'], ' 此人未报名')

    def test_post_activity_registration_success(self):
        openid = self.default_volunteer_attrs['openid']
        with self.app.app_context():
            participant = ActivityParticipant(activity_id=1, participant_openid=openid)
            db.session.add(participant)
            db.session.commit()

        response = self.client.post('/activityRegistration/1?openid={}'.format(openid))
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertEqual(data['status'], 'success')

# GET "/activity/export/sign"
class TestActivityExportSign(ActivityCase):
    def test_export(self):
        headers = {}
        for idx in range(2):
            response = self.client.get(f'/activity/export/sign?activity_id={idx + 1}', headers=headers)
            self.assertEqual(response.status_code, 200)

            workbook = load_workbook(filename=io.BytesIO(response.data))
            ws = workbook['志愿']
            self.assertEqual(ws['A1'].value, '志愿 签到/签收表')
            self.assertTrue(ws['B2'].value.index(self.default_projects[idx]['name']) != -1)

            ws = workbook['视障']
            self.assertEqual(ws['A1'].value, '视障 签到/签收表')

# GET "/activity/export/duty_summary"
class TestExportDutySummary(ActivityCase):
    def test_export(self):
        headers = {}
        today = datetime.now().strftime('%Y-%m-%d')

        response = self.client.get(f'/activity/export/duty_summary?start_date={today}&end_date={today}', headers=headers)
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=io.BytesIO(response.data))
        ws = workbook.active
        self.assertEqual(ws['A1'].value, '活动序号')
        self.assertEqual(ws['B2'].value, today)
        self.assertEqual(ws['C3'].value, self.default_volunteer_attrs['name'])

# GET "/activity/export/info_summary"
class TestExportInfoSummary(ActivityCase):
    def test_export(self):
        headers = {}
        today = datetime.now().strftime('%Y-%m-%d')

        response = self.client.get(f'/activity/export/info_summary?start_date={today}&end_date={today}', headers=headers)
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=io.BytesIO(response.data))
        ws = workbook.active
        self.assertEqual(ws['A2'].value, '活动序号')
        self.assertEqual(ws['B3'].value, today)
        self.assertEqual(ws['E4'].value, 1)

# GET "/activity/projects"
class TestGetProjects(ActivityCase):
    def test_get_all(self):
        headers = {}
        response = self.client.get(f'/activity/projects', headers=headers)
        self.assertEqual(response.status_code, 200)
        data = response.json
        self.assertEqual(data['status'], 'success')
        self.assertEqual(len(data['data']), 2)
        self.assertEqual(data['data'][0]['name'], 'project_1')

# GET "/activity/duties"
class TestGetDuties(ActivityCase):
    def test_get_all(self):
        headers = {}
        response = self.client.get(f'/activity/duties', headers=headers)
        self.assertEqual(response.status_code, 200)
        data = response.json
        self.assertEqual(data['status'], 'success')
        self.assertEqual(len(data['data']), 6)
        self.assertEqual(set([x['name'] for x in data['data']]), set(('志愿者领队', '视障者领队', '主持', '推文写作', '公众号编辑', '拍照')))

# GET "/activity/gifts"
class TestGetGifts(ActivityCase):
    def test_get_all(self):
        headers = {}
        response = self.client.get(f'/activity/gifts', headers=headers)
        self.assertEqual(response.status_code, 200)
        data = response.json
        self.assertEqual(data['status'], 'success')
        self.assertEqual(len(data['data']), 4)
        self.assertEqual(set([x['name'] for x in data['data']]), set(('衣服', '帽子', '臂包', '腰包')))
        self.assertEqual(set([x['category'] for x in data['data']]), set(('other',)))

# POST "/activity/review"
class TestActivityReview(ActivityCase):
    def test_success(self):
        remark = 'I enjoyed it'
        activity_id = 2
        with self.app.app_context():
            participant = db.session.query(ActivityParticipant).filter_by(participant_openid=self.user_helper_attrs[1]['openid']).first()
            self.assertEqual(participant.remark, '')
            self.assertIsNone(participant.duties)
            self.assertIsNone(participant.gifts)
            self.assertIsNone(participant.current_state)

            existing_participant_count = db.session.query(ActivityParticipant).filter_by(activity_id=activity_id).count()

        headers = {'content_type': 'application/json'}
        duties = [1,2]
        gifts = {1: 1, 2: 2}
        response = post_json(self.client, f'/activity/review/{activity_id}',
                data={
                    'data': [{
                        'phone': self.user_helper_attrs[1]['phone'],
                        'duties': duties,
                        'gifts': gifts,
                        'remark': remark,
                        'signup': 1,
                    }, { # new add participant
                        'phone': self.user_helper_attrs[0]['phone'],
                        'duties': duties,
                        'gifts': gifts,
                        'remark': remark,
                        'signup': 0,
                        'is_child': 1,
                    }]
                }, headers=headers)

        with self.app.app_context():
            participants = db.session.query(
                    ActivityParticipant).filter_by(
                    activity_id=activity_id).all()
            self.assertEqual(len(participants), existing_participant_count + 1)
            for participant in participants:
                if participant.user.phone not in [x['phone'] for x in self.user_helper_attrs]:
                    continue

                self.assertEqual(participant.remark, remark)
                self.assertEqual(json.dumps(participant.duties), json.dumps(duties))
                self.assertEqual(json.dumps(participant.gifts), json.dumps(gifts))
                if participant.user.phone == self.user_helper_attrs[1]['phone']:
                    self.assertEqual(participant.current_state, 'signed_up')
                    self.assertEqual(participant.is_child, False)
                if participant.user.phone == self.user_helper_attrs[0]['phone']:
                    self.assertIsNone(participant.current_state)
                    self.assertEqual(participant.is_child, True)
